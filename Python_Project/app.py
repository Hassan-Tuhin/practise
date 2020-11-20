import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def Process_workbook(filename):
    wb=xl.load_workbook(filename)
    sheet=wb["Sheet1"]
    for row in range(4,8):
        cell=sheet.cell(row,4)
        corrected_price=cell.value*2
        corrected_price_cell=sheet.cell(row,5)
        corrected_price_cell.value=corrected_price
    values=Reference(sheet,
            min_row=4,
            max_row=7,
            min_col=4,
            max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename) 